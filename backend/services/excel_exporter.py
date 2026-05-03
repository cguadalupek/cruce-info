from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .constants import RESULT_COLORS, VISIBLE_EXPORT_COLUMNS


def build_export_file(semana: str, rows: list[dict[str, str]]) -> BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Resultado"

    last_column_index = len(VISIBLE_EXPORT_COLUMNS)
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column_index)

    title_cell = worksheet.cell(row=1, column=1, value=f"SEMANA {semana}")
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    header_fill = PatternFill(fill_type="solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    header_row = 2
    data_start_row = 3

    for index, (_, header_label) in enumerate(VISIBLE_EXPORT_COLUMNS, start=1):
        cell = worksheet.cell(row=header_row, column=index, value=header_label)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row_index, row in enumerate(rows, start=data_start_row):
        result_color = RESULT_COLORS.get(row.get("resultado_cruce", ""), "FFFFFF")
        data_fill = PatternFill(fill_type="solid", fgColor=result_color)

        for column_index, (field_name, _) in enumerate(VISIBLE_EXPORT_COLUMNS, start=1):
            cell = worksheet.cell(row=row_index, column=column_index, value=row.get(field_name, ""))
            cell.fill = data_fill
            cell.border = border
            if field_name == "fecha":
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="top")

    for column_index, (field_name, header_label) in enumerate(VISIBLE_EXPORT_COLUMNS, start=1):
        max_length = len(header_label)
        for row in rows:
            max_length = max(max_length, len(str(row.get(field_name, "") or "")))
        worksheet.column_dimensions[worksheet.cell(row=header_row, column=column_index).column_letter].width = min(
            max(max_length + 2, 14),
            45,
        )

    worksheet.freeze_panes = "A3"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
